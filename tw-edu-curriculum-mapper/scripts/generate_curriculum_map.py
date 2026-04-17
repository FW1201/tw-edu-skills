#!/usr/bin/env python3
"""課程地圖 Excel 生成腳本"""
import argparse
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 色彩
C_DEEP_BLUE  = "1A5276"
C_MID_BLUE   = "2471A3"
C_LIGHT_BLUE = "EBF5FB"
C_ALT_ROW    = "F8F9FA"
C_HEADER_TXT = "FFFFFF"
C_GOLD       = "D4AC0D"
C_GREEN      = "1E8449"
C_ORANGE     = "CA6F1E"
C_GRAY       = "566573"

def hdr_font(size=11, bold=True, color=C_HEADER_TXT):
    return Font(name='微軟正黑體', size=size, bold=bold, color=color)

def body_font(size=11, bold=False, color="1C2A35"):
    return Font(name='微軟正黑體', size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color="2471A3")
    return Border(left=s, right=s, top=s, bottom=s)

def center_align(wrap=True):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

def style_header(cell, text, bg=C_MID_BLUE, size=11):
    cell.value = text
    cell.font = hdr_font(size=size)
    cell.fill = fill(bg)
    cell.alignment = center_align()
    cell.border = thin_border()

def style_data(cell, text, row_idx=0, center=False):
    cell.value = text
    bg = C_LIGHT_BLUE if row_idx % 2 == 0 else C_ALT_ROW
    cell.fill = fill(bg)
    cell.font = body_font()
    cell.alignment = center_align() if center else left_align()
    cell.border = thin_border()


def create_overview_sheet(wb, subject, grade, semester, units):
    ws = wb.active
    ws.title = "學期課程總覽"

    # 標題列
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{grade}｜{subject}｜{semester}學期課程地圖"
    title_cell.font = Font(name='微軟正黑體', size=16, bold=True, color=C_DEEP_BLUE)
    title_cell.fill = fill("D6EAF8")
    title_cell.alignment = center_align()
    ws.row_dimensions[1].height = 36

    # 表頭
    headers = ["週次", "單元名稱", "教學主題/重點", "節數", "核心素養", "學習表現", "評量方式", "重大議題"]
    col_widths = [6, 18, 28, 6, 22, 22, 16, 14]
    for j, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=j)
        style_header(cell, h, bg=C_DEEP_BLUE)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 22

    # 資料列（範例）
    week = 1
    for i, unit in enumerate(units):
        row = i + 3
        sample_data = [
            f"{week}–{week+1}",
            unit,
            f"《{unit}》核心概念與技能教學",
            "4",
            "語-J-B1：閱讀理解與表達",
            "5-Ⅳ-2：分析篇章結構",
            "學習單、口頭問答",
            "品德教育（品）",
        ]
        for j, text in enumerate(sample_data, 1):
            style_data(ws.cell(row=row, column=j), text, row_idx=i,
                       center=(j in [1, 4]))
        ws.row_dimensions[row].height = 30
        week += 2

    # 凍結首兩列
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{2 + len(units)}"


def create_performance_sheet(wb, units):
    ws = wb.create_sheet("學習表現對應")
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "學習表現對應表"
    c.font = Font(name='微軟正黑體', size=14, bold=True, color=C_DEEP_BLUE)
    c.fill = fill("D6EAF8")
    c.alignment = center_align()
    ws.row_dimensions[1].height = 32

    headers = ["單元名稱", "學習表現代碼", "說明", "布魯姆層次", "評量方式", "備註"]
    col_widths = [18, 14, 32, 12, 18, 14]
    for j, (h, w) in enumerate(zip(headers, col_widths), 1):
        style_header(ws.cell(row=2, column=j), h, bg=C_MID_BLUE)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 22

    bloom_levels = ["C1 記憶", "C2 理解", "C3 應用", "C4 分析", "C5 評鑑", "C6 創造"]
    sample_perf = ["5-Ⅳ-1", "5-Ⅳ-2", "6-Ⅳ-1"]
    sample_desc = ["閱讀並理解多元文本", "分析篇章結構與寫作手法", "依文體撰寫作文"]

    row = 3
    for i, unit in enumerate(units):
        for k in range(min(2, len(sample_perf))):
            data = [unit if k == 0 else "", sample_perf[k % 3],
                    sample_desc[k % 3], bloom_levels[k+1],
                    "學習單、口頭", ""]
            for j, text in enumerate(data, 1):
                style_data(ws.cell(row=row, column=j), text, row_idx=i)
            ws.row_dimensions[row].height = 22
            row += 1

    ws.freeze_panes = "A3"


def create_content_sheet(wb, units):
    ws = wb.create_sheet("學習內容對應")
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "學習內容對應表"
    c.font = Font(name='微軟正黑體', size=14, bold=True, color=C_DEEP_BLUE)
    c.fill = fill("D6EAF8")
    c.alignment = center_align()
    ws.row_dimensions[1].height = 32

    headers = ["單元名稱", "學習內容代碼", "說明", "教學材料/媒材", "備註"]
    col_widths = [18, 14, 32, 22, 14]
    for j, (h, w) in enumerate(zip(headers, col_widths), 1):
        style_header(ws.cell(row=2, column=j), h, bg=C_MID_BLUE)
        ws.column_dimensions[get_column_letter(j)].width = w

    sample_cont = [("Ac-Ⅳ-3", "記敘文的時間、空間、人物情節"),
                   ("Ca-Ⅳ-1", "修辭技巧運用")]
    row = 3
    for i, unit in enumerate(units):
        for k, (code, desc) in enumerate(sample_cont):
            data = [unit if k == 0 else "", code, desc, "教科書、補充教材", ""]
            for j, text in enumerate(data, 1):
                style_data(ws.cell(row=row, column=j), text, row_idx=i)
            ws.row_dimensions[row].height = 22
            row += 1
    ws.freeze_panes = "A3"


def create_crossdomain_sheet(wb, units, subject):
    ws = wb.create_sheet("跨域連結矩陣")
    related = ["英語", "社會", "自然", "藝術", "健體", "資訊", "綜合"]

    ws.merge_cells(f"A1:{get_column_letter(len(related)+1)}1")
    c = ws["A1"]
    c.value = f"{subject}跨域連結矩陣"
    c.font = Font(name='微軟正黑體', size=14, bold=True, color=C_DEEP_BLUE)
    c.fill = fill("D6EAF8")
    c.alignment = center_align()
    ws.row_dimensions[1].height = 32

    # 表頭
    style_header(ws.cell(row=2, column=1), "單元名稱", bg=C_DEEP_BLUE)
    ws.column_dimensions["A"].width = 20
    for j, subj in enumerate(related, 2):
        style_header(ws.cell(row=2, column=j), subj, bg=C_DEEP_BLUE)
        ws.column_dimensions[get_column_letter(j)].width = 10
    ws.row_dimensions[2].height = 22

    symbols = {"強": ("●", "1A5276"), "中": ("◎", "D4AC0D"), "": ("", C_ALT_ROW)}
    import random
    random.seed(42)
    for i, unit in enumerate(units):
        row = i + 3
        style_data(ws.cell(row=row, column=1), unit, row_idx=i)
        for j in range(2, len(related) + 2):
            level = random.choice(["強", "中", "中", ""])
            sym, color = symbols.get(level, ("", C_ALT_ROW))
            cell = ws.cell(row=row, column=j)
            bg = C_LIGHT_BLUE if i % 2 == 0 else C_ALT_ROW
            cell.value = sym
            cell.fill = fill(bg)
            cell.font = Font(name='微軟正黑體', size=12, bold=(level=="強"), color=color)
            cell.alignment = center_align()
            cell.border = thin_border()
        ws.row_dimensions[row].height = 22

    # 圖例
    legend_row = len(units) + 4
    ws.cell(row=legend_row, column=1).value = "圖例："
    ws.cell(row=legend_row, column=1).font = body_font(bold=True)
    ws.cell(row=legend_row, column=2).value = "● 強連結"
    ws.cell(row=legend_row, column=2).font = Font(name='微軟正黑體', size=11, color="1A5276")
    ws.cell(row=legend_row, column=3).value = "◎ 中等連結"
    ws.cell(row=legend_row, column=3).font = Font(name='微軟正黑體', size=11, color="D4AC0D")

    ws.freeze_panes = "B3"


def create_detail_sheet(wb, subject, grade, semester, units):
    ws = wb.create_sheet("教學活動詳表")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"{grade}｜{subject}｜{semester}學期教學活動詳表"
    c.font = Font(name='微軟正黑體', size=15, bold=True, color=C_DEEP_BLUE)
    c.fill = fill("D6EAF8")
    c.alignment = center_align()
    ws.row_dimensions[1].height = 34

    headers = ["單元名稱", "階段", "建議時間", "教學目標", "教師活動", "學生任務", "引導問題", "媒材", "評量方式", "備註"]
    col_widths = [18, 10, 10, 18, 28, 22, 22, 18, 14, 14]
    for j, (h, w) in enumerate(zip(headers, col_widths), 1):
        style_header(ws.cell(row=2, column=j), h, bg=C_DEEP_BLUE)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 24

    phases = [
        ("導入一", "4分",
         "{unit}導入情境與先備知識激活",
         "以 {unit} 的生活情境或核心畫面開場，先讓學生說出自己的觀察，再把回答整理成今天要學的關鍵詞。",
         "學生先口頭預測主題，再寫下與 {unit} 有關的既有經驗，建立共同背景。",
         "你看到什麼？你想到什麼？",
         "情境圖、封面、關鍵字卡",
         "口頭問答、快速筆記",
         "開場一定要和學生經驗連結"),
        ("導入二", "4分",
         "{unit}學習目標與任務說明",
         "清楚說明本節課為什麼要學 {unit}，要學生知道今天不是只看內容，而是要學會怎麼分析與怎麼表達。",
         "學生複述學習目標，並選擇自己最想弄懂的一個問題。",
         "今天學完之後，你希望自己能做到什麼？",
         "目標投影片、任務卡",
         "口頭複述、任務確認",
         "把學習目的說明白"),
        ("發展一", "8分",
         "{unit}主概念拆解與示範",
         "教師逐步拆解 {unit} 的主概念，示範如何找出重點、標註關鍵句、圈出例證，並口語化說明推理過程。",
         "學生跟著教師進行標註，完成第一輪理解，並記下自己最不確定的部分。",
         "這一句為什麼重要？它和前一句有什麼關係？",
         "投影片、板書、文本/題目",
         "觀察、圈選、隨堂提問",
         "把分析過程做給學生看"),
        ("發展二", "8分",
         "{unit}範例操作與小組討論",
         "用一個完整範例帶學生操作，把 {unit} 的分析方法轉成可實作步驟，並安排小組討論來驗證理解是否一致。",
         "學生分組說明自己的答案，補充理由，並互相修正答案中的模糊處。",
         "如果答案不同，你怎麼證明自己的推論？",
         "範例講義、學習單、小白板",
         "小組發表、同儕互評",
         "讓學生真的做一次"),
        ("應用一", "6分",
         "{unit}獨立練習與即時回饋",
         "提供一個和 {unit} 相似但不完全相同的練習，讓學生獨立作答，教師在旁巡迴給回饋與提示。",
         "學生獨立完成練習，並用一到兩句話說明自己的判斷依據。",
         "你是根據哪個線索做出判斷？",
         "練習題、作答紙、計時器",
         "紙本作答、巡迴觀察",
         "確認能不能自己做"),
        ("總結一", "3分",
         "{unit}重點收束與學習檢核",
         "把 {unit} 的重點重新整理成 3 個可帶走的關鍵句，並提醒學生下一次上課或作業要接續什麼內容。",
         "學生完成出口票，寫出今天學到的三件事與一個仍然不清楚的問題。",
         "今天最重要的三個重點是什麼？",
         "出口票、回饋卡、摘要條",
         "出口票、口頭回饋",
         "用簡短方式收尾"),
        ("延伸一", "3分",
         "{unit}延伸應用與跨域連結",
         "把 {unit} 延伸到日常生活、跨領域主題或下一份作業，讓學生知道這個概念不只存在課堂裡。",
         "學生把學到的內容和自己的生活經驗或其他學科連結起來。",
         "這個概念還能用在哪裡？",
         "延伸閱讀、生活情境卡",
         "口頭分享、延伸作業",
         "留下下一步"),
    ]

    row = 3
    for i, unit in enumerate(units):
        for phase_name, duration, goal, teacher, student, question, media, assess, note in phases:
            values = [
                unit,
                phase_name,
                duration,
                goal.format(unit=unit, subject=subject, grade=grade, semester=semester),
                teacher.format(unit=unit, subject=subject, grade=grade, semester=semester),
                student.format(unit=unit, subject=subject, grade=grade, semester=semester),
                question.format(unit=unit, subject=subject, grade=grade, semester=semester),
                media,
                assess,
                note,
            ]
            for j, text in enumerate(values, 1):
                style_data(ws.cell(row=row, column=j), text, row_idx=i)
            ws.row_dimensions[row].height = 38
            row += 1

    ws.freeze_panes = "A3"


def create_script_sheet(wb, subject, grade, semester, units):
    ws = wb.create_sheet("教學腳本全文")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{grade}｜{subject}｜{semester}學期教學腳本全文"
    c.font = Font(name='微軟正黑體', size=15, bold=True, color=C_DEEP_BLUE)
    c.fill = fill("D6EAF8")
    c.alignment = center_align()
    ws.row_dimensions[1].height = 34

    headers = [
        "單元名稱", "起始引導", "核心講解", "教師示範", "學生練習",
        "概念難點", "常見誤區", "評量設計", "差異化支持", "延伸作業",
        "收束提醒", "備註",
    ]
    col_widths = [18, 28, 28, 28, 28, 24, 24, 24, 24, 24, 24, 14]
    for j, (h, w) in enumerate(zip(headers, col_widths), 1):
        style_header(ws.cell(row=2, column=j), h, bg=C_MID_BLUE)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 24

    for i, unit in enumerate(units):
        intro = (
            f"在 {grade} 的 {subject} 課堂中，以《{unit}》作為課程切入點，先用一句能喚起興趣的問題帶出今天的任務，"
            f"讓學生知道這一課不只是理解內容，而是要學會如何抓重點、如何說明理由、以及如何把 {unit} 的概念轉成自己的語言。"
            f"教師可先以生活情境、圖片、短句或前導題建立共鳴，並在黑板上寫下三個關鍵字，讓全班快速進入同一個理解框架。"
        )
        core = (
            f"接著進入 {unit} 的核心概念講解，教師需要把抽象內容拆成可操作步驟：先辨識主要訊息，再找出支持證據，"
            f"然後將它們連成完整推理。若是語文文本，就示範如何標註情感詞、關鍵句與結構；若是概念型單元，就示範如何分類、比較與歸納。"
            f"此階段的講解不能只講答案，要讓學生看見思考過程，知道為什麼這樣判斷，以及不同答案之間差在哪裡。"
        )
        demo = (
            f"教師示範部分可用 {unit} 的一小段內容作為實作範例，當場把分析步驟做出來，讓學生看到完整過程："
            f"先圈出題幹或文本中的關鍵詞，再標出限制條件與提示線索，最後用一句話說明結論。"
            f"若班上學生程度差異大，可一邊示範一邊補充簡化版與進階版做法，讓不同程度的學生都能跟上。"
        )
        practice = (
            f"學生練習時，安排一個與《{unit}》相近但不完全相同的題目，要求學生自己完成一次判斷，"
            f"並寫下做出答案的依據。教師巡迴時要即時提供回饋，尤其注意學生是否只是背答案，還是能夠說明理由。"
            f"若需要，可先讓學生兩人一組討論，再轉成個人作答，讓理解從口語化的互動逐漸收斂到可檢核的書面結果。"
        )
        difficulty = (
            f"《{unit}》最容易卡住的地方通常是概念抽象或步驟過快，所以要先將難點拆成短句、圖示或對照例，"
            f"避免學生在第一輪就失去方向。若是文本理解，就要處理人物關係、事件順序或象徵意義；若是概念題，就要處理條件、公式、分類標準或判斷依據。"
        )
        misconception = (
            f"常見誤區包含：把表面答案當成完整理解、只會背定義卻不會舉例、或在 {unit} 的分析中忽略前後文與證據。"
            f"因此教師需要用反例、比較表與口語追問，幫學生看見自己到底是理解了概念，還是只是在猜答案。"
        )
        assessment = (
            f"評量設計建議採形成性為主：可用口頭問答、出口票、短寫作、簡答題或小組發表來檢核 {unit} 的掌握程度。"
            f"每個評量點都應對應一個明確學習目標，讓老師能快速看出學生是理解、應用、還是仍停留在記憶層次。"
        )
        differentiation = (
            f"差異化支持可以分成三層：基礎組提供關鍵詞提示與步驟卡，標準組完成完整分析，進階組則加入延伸推理或跨域比較。"
            f"這樣做能讓不同程度的學生都在同一節課中前進，而不是只有少數學生跟上。"
        )
        homework = (
            f"延伸作業可以請學生把 {unit} 的方法用到另一個文本、另一題題目或生活情境中，"
            f"讓學習不只停在當堂答案，而是能夠被遷移到新的脈絡，形成真正可重複使用的能力。"
        )
        close = (
            f"收束時，教師要把《{unit}》的學習重點壓縮成三句話，提醒學生今天學會的方法可以用在下一課、下一章或其他學科。"
            f"最後用出口票或一句摘要檢查學習成效，並告訴學生下次上課前可以先預習什麼，讓課程之間保持連續感。"
            f"這一段也適合加入延伸閱讀、生活應用或跨域連結，讓知識不只停留在課堂上。"
        )
        note = "注意節奏與閱讀量，避免單頁過密。"
        values = [unit, intro, core, demo, practice, difficulty, misconception, assessment, differentiation, homework, close, note]
        row = i + 3
        for j, text in enumerate(values, 1):
            style_data(ws.cell(row=row, column=j), text, row_idx=i)
            ws.cell(row=row, column=j).alignment = left_align(wrap=True)
        ws.row_dimensions[row].height = 110

    ws.freeze_panes = "A3"


def create_teacher_memo_sheet(wb, subject, grade, semester, units):
    ws = wb.create_sheet("教師備忘錄")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{grade}｜{subject}｜{semester}學期教師備忘錄"
    c.font = Font(name='微軟正黑體', size=15, bold=True, color=C_DEEP_BLUE)
    c.fill = fill("D6EAF8")
    c.alignment = center_align()
    ws.row_dimensions[1].height = 34

    headers = ["單元名稱", "教學焦點", "常見提問", "分層提示", "補充資源", "教學警示", "下一步建議", "備註"]
    col_widths = [18, 28, 28, 28, 28, 28, 28, 14]
    for j, (h, w) in enumerate(zip(headers, col_widths), 1):
        style_header(ws.cell(row=2, column=j), h, bg=C_DEEP_BLUE)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 24

    for i, unit in enumerate(units):
        focus = (
            f"《{unit}》的教學焦點是讓學生從表面理解走向結構理解，因此教師要反覆提醒：今天不是只找答案，"
            f"而是要看懂推理的路徑。若是語文文本，焦點放在情感、結構與修辭；若是概念單元，焦點放在分類、比較與應用。"
        )
        questions = (
            f"上課時可以常問：這裡最重要的線索是什麼？如果你換一個例子還成立嗎？你會如何用自己的話說明？"
            f"這些問題能幫學生從記憶性回答轉向理解性回答，也能看出誰真正掌握了 {unit} 的方法。"
        )
        support = (
            f"基礎學生可提供關鍵詞或句型框架；中等學生可要求自己完成完整推論；進階學生則加上比較、反證或跨域遷移。"
            f"分層不是降低標準，而是讓不同起點的學生都能朝同一個理解目標前進。"
        )
        resources = (
            f"補充資源可包含課文原文、簡化版講義、圖像提示卡、範例答案與延伸閱讀。"
            f"若班級需要更多視覺支援，可把 {unit} 的關鍵步驟做成流程圖或對照表。"
        )
        warning = (
            f"最常見的風險是學生只記住結論而忽略理由，或是在 {unit} 的練習中被過多資訊分散注意力。"
            f"教師要避免一次塞入太多內容，並在每個階段都保留回問與停頓。"
        )
        next_step = (
            f"下一步建議把 {unit} 的方法套用到另一段文本或題目，讓學生在新情境中再次使用同一套思考流程，"
            f"這樣才算真的把概念學會，而不是只會在單一題目上作答。"
        )
        note = "供教師備課時快速掃描。"
        values = [unit, focus, questions, support, resources, warning, next_step, note]
        row = i + 3
        for j, text in enumerate(values, 1):
            style_data(ws.cell(row=row, column=j), text, row_idx=i)
            ws.cell(row=row, column=j).alignment = left_align(wrap=True)
        ws.row_dimensions[row].height = 96

    ws.freeze_panes = "A3"


def create_appendix_sheet(wb, subject, grade, semester, units):
    ws = wb.create_sheet("課程補充說明")

    ws.merge_cells("A1:A1")
    c = ws["A1"]
    c.value = f"{grade}｜{subject}｜{semester}學期補充說明"
    c.font = Font(name='微軟正黑體', size=15, bold=True, color=C_DEEP_BLUE)
    c.fill = fill("D6EAF8")
    c.alignment = center_align()
    ws.row_dimensions[1].height = 34

    ws["A3"] = "使用方式"
    ws["A3"].font = hdr_font(size=12)
    ws["A3"].fill = fill(C_DEEP_BLUE)
    ws["A3"].alignment = center_align()
    ws["A3"].border = thin_border()
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90

    note_rows = [
        f"《{units[0] if units else '單元'}》建議先從文本脈絡與核心概念切入，讓學生知道這一單元的學習重點不只是答題，而是要學會怎麼看懂資料、怎麼表達理由、以及怎麼把概念轉成自己的語言。",
        f"在 {subject} 課程中，{grade} 的學生特別需要清楚的步驟提示，因此每個單元都要保留示範、練習、回饋三段。若 {semester} 學期的課程進度較緊，也至少保留一個形成性檢核與一個延伸任務。",
        f"這份課程地圖不是只做給行政或備課檢查，而是要讓教師在實際教學時能快速看見單元彼此之間的節奏、難度與跨域連結。若某單元學生反應不佳，可直接回到此表調整節數、素材與評量方式。",
    ]

    for idx, text in enumerate(note_rows, 4):
        cell = ws.cell(row=idx, column=2)
        cell.value = text
        cell.font = body_font(size=11)
        cell.fill = fill(C_LIGHT_BLUE if idx % 2 == 0 else C_ALT_ROW)
        cell.alignment = left_align(wrap=True)
        cell.border = thin_border()
        ws.row_dimensions[idx].height = 54
        ws.cell(row=idx, column=1).value = f"說明 {idx-3}"
        ws.cell(row=idx, column=1).font = body_font(bold=True)
        ws.cell(row=idx, column=1).fill = fill(C_LIGHT_BLUE if idx % 2 == 0 else C_ALT_ROW)
        ws.cell(row=idx, column=1).alignment = center_align()
        ws.cell(row=idx, column=1).border = thin_border()

    ws.freeze_panes = "A3"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--subject',  default='國語文')
    parser.add_argument('--grade',    default='國中八年級')
    parser.add_argument('--semester', default='上')
    parser.add_argument('--units',    default='第一課,第二課,第三課,第四課,第五課,第六課')
    parser.add_argument('--output',   default='課程地圖.xlsx')
    args = parser.parse_args()

    units = [u.strip() for u in args.units.split(',') if u.strip()]

    wb = Workbook()
    create_overview_sheet(wb, args.subject, args.grade, args.semester, units)
    create_performance_sheet(wb, units)
    create_content_sheet(wb, units)
    create_detail_sheet(wb, args.subject, args.grade, args.semester, units)
    create_script_sheet(wb, args.subject, args.grade, args.semester, units)
    create_teacher_memo_sheet(wb, args.subject, args.grade, args.semester, units)
    create_appendix_sheet(wb, args.subject, args.grade, args.semester, units)
    create_crossdomain_sheet(wb, units, args.subject)

    wb.save(args.output)
    print(f"✓ 課程地圖已儲存：{args.output}")

if __name__ == '__main__':
    main()
